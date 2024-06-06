import argparse
import json
import re
import sys
from pymongo import MongoClient
import subprocess
from openpyxl import Workbook
import os
from openpyxl.drawing.image import Image
import requests
from frameioclient import FrameioClient
#not using pandas so i need a good grade lmao
API_TOKEN = "fio-u-7SxiUam_bwKNxyXSafmnMNXhxnfFOMFPGLaYuwVzypQcs4UjsruJ2N4mcbdKEjaU"
FOLDER_ID = "e7af95ba-329e-4694-8af0-0974415cf526"

def parse_args():
    parser = argparse.ArgumentParser(description="Process Baselight and Xytech files and populate database.")
    parser.add_argument("--baselight", default='C:\\Users\\yungj\\OneDrive\\Desktop\\467 CHAJA\\Project1467\\Baselight_export.txt', help="Path to Baselight file")
    parser.add_argument("--xytech", default='C:\\Users\\yungj\\OneDrive\\Desktop\\467 CHAJA\\Project1467\\Xytech.txt', help="Path to Xytech file")
    parser.add_argument("--process", help="make the file process the video")
    parser.add_argument("--output", required=False, help="output Excel file path")
    parser.add_argument("--send", action='store_true', help="send the videos to frame.io")
    return parser.parse_args()

def parse_xytech(file_path):
    with open(file_path, 'r') as file:
        content = file.read()
    producer = re.search(r"Producer:\s*(.+)", content).group(1)
    operator = re.search(r"Operator:\s*(.+)", content).group(1)
    job = re.search(r"Job:\s*(.+)", content).group(1)
    notes = re.search(r"Notes:\s*(.+)", content, re.DOTALL).group(1).strip()
    return producer, operator, job, notes

def parse_baselight(file_path):
    frame_data = []
    hpsans_mapping = {
        'VFX/Hydraulx': '/hpsans12/production',
        'VFX/Framestore': '/hpsans13/production',
        'VFX/AnimalLogic': '/hpsans14/production',
        'shot_1ab': '/hpsans15/production',
        'shot_2b': '/hpsans11/production',
        'partC': '/hpsans17/production',
    }
    with open(file_path, 'r') as file:
        lines = file.readlines()
    for line in lines:
        if not line.strip() or '/baselightfilesystem1' not in line:
            continue
        path = line.strip()
        replaced = False
        for key, value in hpsans_mapping.items():
            if key in path:
                path = path.replace('/baselightfilesystem1', value)
                replaced = True
                break
        if not replaced:  #if not apply determined
            path = path.replace('/baselightfilesystem1', '/hpsans13/production')

        # sort the frames
        path, *frames = path.split()
        frames = sorted(int(frame) for frame in frames if frame.isdigit())
        organized_frames = organize_into_ranges(frames)
        
        # modified path and organized frames
        frame_data.append([path, organized_frames])
    return frame_data

def organize_into_ranges(frames):
    if not frames:
        return ""
    ranges = []
    start = end = frames[0]
    for frame in frames[1:]:
        if frame == end + 1:
            end = frame
        else:
            ranges.append(f"{start}-{end}" if start != end else str(start))
            start = end = frame
    ranges.append(f"{start}-{end}" if start != end else str(start))
    return ranges

def populate_database(xytech_data, baselight_data):
    client = MongoClient("mongodb://localhost:27017/")
    db = client["project1_db"]
    xytech_collection = db["xytech"]
    baselight_collection = db["baselight"]
    xytech_collection.insert_one({"producer": xytech_data[0], "operator": xytech_data[1], "job": xytech_data[2], "notes": xytech_data[3]})
    for path, ranges in baselight_data:
        baselight_collection.insert_one({"folder": path, "frames": ranges})
#stderr is for my own understanding since I fucked up a lot on what errors i was getting lmao
def get_video_duration(video_path):
    cmd = ['ffprobe', '-v', 'error', '-select_streams', 'v:0', '-show_entries', 'format=duration', '-of', 'default=noprint_wrappers=1:nokey=1', video_path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result.returncode != 0:
        print("error getting video duration: ", result.stderr)
        return None
    try:
        return float(result.stdout.strip())
    except ValueError:
        print("error parsing")
        return None

def get_video_frame_rate(video_path):
    cmd = ['ffprobe', '-v', 'error', '-select_streams', 'v:0', '-show_entries', 'stream=r_frame_rate', '-of', 'json', video_path]
    result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    if result.returncode != 0:
        print("error getting frame rate:", result.stderr)
        return None
    try:
        frame_rate_info = json.loads(result.stdout)
        frame_rate_str = frame_rate_info['streams'][0]['r_frame_rate']
        num, den = map(int, frame_rate_str.split('/'))
        return num / den
    except (KeyError, IndexError, ValueError):
        print("error parsing frame rate.")
        return None

def frame_to_timecode(frame_number, frame_rate):
    total_seconds = frame_number / frame_rate  #float division for timecode 
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    frame_remainder = int(round((total_seconds - int(total_seconds)) * frame_rate))

    return f"{hours:02}:{minutes:02}:{seconds:02}:{frame_remainder:02}"

def frame_to_timecode_ms(frame_number, frame_rate):
    total_seconds = frame_number / frame_rate  # i tried using this for the videos in general instead of the general format HH:MM:SS:FF and switch last frame remainder to ML
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    milliseconds = int((total_seconds - int(total_seconds)) * 1000)  #  fraction to milliseconds

    return f"{hours:02}:{minutes:02}:{seconds:02}.{milliseconds:03}"

def create_thumbnail(video_path, frame_number, output_path):
    command = [
        'ffmpeg', '-y',  
        '-i', video_path,
        '-vf', f'select=eq(n\,{frame_number})',
        '-vframes', '1',
        '-s', '96x74',  # image to 96x74
        output_path
    ]
    subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

def create_video_clip(video_path, start_frame, end_frame, frame_rate, output_path):
    start_timecode = frame_to_timecode_ms(start_frame, frame_rate)
    end_timecode = frame_to_timecode_ms(end_frame, frame_rate)
    command = [
        'ffmpeg', '-y',  
        '-i', video_path,
        '-ss', start_timecode,
        '-to', end_timecode,
        output_path
    ]
    subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

def get_middle_frame(start_frame, end_frame):
    return start_frame + (end_frame - start_frame) // 2

def write_to_excel(data, file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Location', 'Frames to Fix', 'Timecodes', 'Thumbnails'])  # headers

    for folder, frame_range, timecodes, thumbnail_path in data:
        row = [folder, frame_range, timecodes]
        ws.append(row)
        if os.path.exists(thumbnail_path):
            img = Image(thumbnail_path)
            img.anchor = ws.cell(row=ws.max_row, column=4).coordinate
            ws.add_image(img)
            
    # column widths to auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"written to Excel at {file_path}")

def fetch_and_print_baselight_data(frame_rate, video_length, video_path, collect_for_excel=False):
    client = MongoClient("mongodb://localhost:27017/")
    db = client["project1-3_db"]
    baselight_collection = db["baselight"]
    entries = baselight_collection.find({})
    excel_data = []

    thumbnail_folder = 'thumbnails'
    os.makedirs(thumbnail_folder, exist_ok=True)

    video_folder = 'videos'  #local file to the folder created with the videos to send them to frame.io
    os.makedirs(video_folder, exist_ok=True)

    max_frame = int(video_length * frame_rate) if video_length is not None else float('inf')

    for entry in entries:
        folder = entry['folder']
        for frame_range in entry['frames']:
            if '-' in frame_range:
                start_frame, end_frame = map(int, frame_range.split('-'))
                if start_frame <= max_frame and end_frame <= max_frame:
                    middle_frame = get_middle_frame(start_frame, end_frame)
                    thumbnail_path = f"{thumbnail_folder}/{frame_range}.jpg"
                    create_thumbnail(video_path, middle_frame, thumbnail_path)
                    
                    start_timecode = frame_to_timecode(start_frame, frame_rate)
                    end_timecode = frame_to_timecode(end_frame, frame_rate)
                    video_clip_path = f"{video_folder}/{frame_range}.mp4"
                    create_video_clip(video_path, start_frame, end_frame, frame_rate, video_clip_path)
                    
                    output = f"{folder}, {frame_range}, {start_timecode} to {end_timecode}"
                    print(output)
                    if collect_for_excel:
                        excel_data.append([folder, frame_range, f"{start_timecode} - {end_timecode}", thumbnail_path])
                else:
                    print(f"skipped: {folder}, {frame_range} as it is beyond the video length.")
            else:
                print(f"skipped: {folder}, {frame_range} as it is a single frame.")

    return excel_data

def upload_videos_to_frame_io():
    video_folder = 'videos'
    if not os.path.exists(video_folder):
        print("folder not found bruh")
        return
    
    client = FrameioClient(API_TOKEN)

    for video_file in os.listdir(video_folder):
        if video_file.endswith(".mp4"):
            video_path = os.path.join(video_folder, video_file)
            try:
                # upload the file
                asset = client.assets.upload(FOLDER_ID, video_path)
                print(f"Uploaded {video_file} successfully.")
            except Exception as e:
                print(f"Failed to send {video_file}. type of error: {e}")

def main():
    args = parse_args()
    
    if args.process:
        xytech_data = parse_xytech(args.xytech)
        baselight_data = parse_baselight(args.baselight)
        populate_database(xytech_data, baselight_data)

        video_duration = get_video_duration(args.process)
        frame_rate = get_video_frame_rate(args.process)
        if video_duration is None or frame_rate is None:
            print("failed to get video information")
            return  
        
        # Ensure the video path is passed correctly
        data_for_excel = fetch_and_print_baselight_data(frame_rate, video_duration, args.process, collect_for_excel=bool(args.output))
        
        if args.output:
            write_to_excel(data_for_excel, args.output)
            print("data successfully written to excel.")
        else:
            print("No output file specified.")
    
    if args.send:
        upload_videos_to_frame_io()

if __name__ == "__main__":
    main()




